#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
コード表.xlsx からウクレレ(4弦)のコードダイアグラム画像を自動生成するスクリプト。

仕組み:
  - Excelの各シート(ルート音ごと)には複数のコードブロックが並んでいる。
  - 各ブロックは「コード名セル」+「フレット番号行(1,2,3,4,5)」+ グリッド枠(4弦x5フレット)で構成。
  - 実際の押弦位置(○)・セーハ(縦長の角丸長方形)は、xlsxパッケージ内の
    xl/drawings/drawingN.xml に図形(ellipse / flowChartTerminator)として
    セル座標(col/row + EMUオフセット)で記録されている。
  - openpyxlはオートシェイプを読めないため、drawing XMLを直接パースする。

出力:
  - chords/<コード名>.png を生成(ファイル名はOSで安全な文字に変換)
"""

import os
import re
import json
import zipfile
import shutil
import unicodedata
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont

XLSX_PATH = os.environ.get("CHORD_XLSX", "コード表.xlsx")
OUT_DIR = os.environ.get("CHORD_OUT_DIR", "docs/chords")
WORK_DIR = "_xlsx_extract_tmp"

NUM_STRINGS = 4   # ウクレレ
NUM_FRETS = 5     # 表示するフレット数

# コード名タイトルのフォントサイズ(全コード共通の固定値)。
# Excel側のコード名フォントは HGPゴシックE 48pt で全コード共通、
# フレット1個分の実寸は約49px(Excel実測値)。この比率(48pt/49px ≈ 1.306)を
# 維持して、フレット幅24pxに対応するフォントサイズを算出している。
TITLE_FONT_SIZE = 31


# ----------------------------------------------------------------------------
# 1. xlsxを展開してdrawing XMLを取得
# ----------------------------------------------------------------------------
def extract_xlsx(xlsx_path, work_dir):
    if os.path.exists(work_dir):
        shutil.rmtree(work_dir)
    os.makedirs(work_dir, exist_ok=True)
    with zipfile.ZipFile(xlsx_path) as z:
        z.extractall(work_dir)
    return work_dir


def get_sheet_to_drawing_map(work_dir, sheet_names_in_order):
    """workbook.xmlのsheet順とsheetN.xmlの対応、さらにdrawing relsを辿る"""
    # workbook.xml.rels: rId -> sheetN.xml
    with open(os.path.join(work_dir, "xl", "_rels", "workbook.xml.rels"), encoding="utf-8") as f:
        wb_rels = f.read()
    rid_to_target = dict(re.findall(r'Id="(rId\d+)"[^>]*Target="([^"]+)"', wb_rels))

    with open(os.path.join(work_dir, "xl", "workbook.xml"), encoding="utf-8") as f:
        wb_xml = f.read()
    # <sheet name="C" sheetId="1" r:id="rId1"/>
    sheet_entries = re.findall(r'<sheet name="([^"]+)"[^>]*r:id="(rId\d+)"', wb_xml)

    name_to_sheetfile = {}
    for name, rid in sheet_entries:
        target = rid_to_target.get(rid)
        if target:
            target = target.replace("worksheets/", "")
            name_to_sheetfile[name] = target  # e.g. sheet1.xml

    name_to_drawing = {}
    for name, sheetfile in name_to_sheetfile.items():
        rels_path = os.path.join(work_dir, "xl", "worksheets", "_rels", sheetfile + ".rels")
        if not os.path.exists(rels_path):
            continue
        with open(rels_path, encoding="utf-8") as f:
            rels = f.read()
        m = re.search(r'Type="[^"]*drawing"[^>]*Target="([^"]+)"', rels)
        if m:
            drawing_target = m.group(1).replace("../drawings/", "")
            name_to_drawing[name] = drawing_target  # e.g. drawing1.xml
    return name_to_drawing


# ----------------------------------------------------------------------------
# 2. drawing XMLから図形(ellipse / flowChartTerminator)を抽出
# ----------------------------------------------------------------------------
def parse_shapes(drawing_path):
    with open(drawing_path, encoding="utf-8") as f:
        content = f.read()

    shapes = []
    anchors = re.findall(r"<xdr:twoCellAnchor[^>]*>(.*?)</xdr:twoCellAnchor>", content, re.S)
    for block in anchors:
        m = re.search(
            r"<xdr:from><xdr:col>(\d+)</xdr:col><xdr:colOff>(\d+)</xdr:colOff>"
            r"<xdr:row>(\d+)</xdr:row><xdr:rowOff>(\d+)</xdr:rowOff></xdr:from>"
            r"<xdr:to><xdr:col>(\d+)</xdr:col><xdr:colOff>(\d+)</xdr:colOff>"
            r"<xdr:row>(\d+)</xdr:row><xdr:rowOff>(\d+)</xdr:rowOff></xdr:to>",
            block,
        )
        if not m:
            continue
        fc, fco, fr, fro, tc, tco, tr, tro = map(int, m.groups())
        prst_m = re.search(r'prst="(\w+)"', block)
        prst = prst_m.group(1) if prst_m else None
        if prst not in ("ellipse", "flowChartTerminator"):
            continue
        rot = bool(re.search(r'<a:xfrm rot="\d+"', block))
        shapes.append(
            {
                "type": prst,
                "from_col": fc,
                "from_row": fr,
                "to_col": tc,
                "to_row": tr,
                "rotated": rot,
            }
        )
    return shapes


# ----------------------------------------------------------------------------
# 3. ワークシートからコードブロック(コード名 + グリッド位置)を検出
# ----------------------------------------------------------------------------
def find_chord_blocks(ws):
    blocks = []
    max_row = ws.max_row
    max_col = ws.max_column

    for row in range(1, max_row + 1):
        for col in range(1, max_col - NUM_FRETS + 2):
            vals = [ws.cell(row=row, column=col + i).value for i in range(NUM_FRETS)]
            # 「1,2,3,4,5」だけでなく、ハイポジション用の「10,11,12,13,14」のような
            # 任意開始値の連続する整数列もフレット数字行として認識する。
            if (
                all(isinstance(v, int) for v in vals)
                and vals == list(range(vals[0], vals[0] + NUM_FRETS))
                and vals[0] >= 1
            ):
                start_fret = vals[0]

                # コード名を探す(フレット行の上、同じ列付近)
                name = None
                for r2 in range(row - 1, max(row - 8, 0), -1):
                    for c2 in (col, col - 1, col + 1, col - 2, col + 2):
                        if c2 < 1:
                            continue
                        v = ws.cell(row=r2, column=c2).value
                        if isinstance(v, str) and v.strip():
                            name = v.strip()
                            break
                    if name:
                        break
                if name is None:
                    continue  # 空テンプレ枠はスキップ

                # グリッド本体はフレット数字行の「上」にある。
                # 罫線は NUM_STRINGS行分(弦と弦の境界線がNUM_STRINGS本)、
                # フレット数字行の直前の行までがグリッドの下端境界。
                grid_row1_top = row - (NUM_STRINGS + 1)  # 1-indexed: グリッド最上端の行

                blocks.append(
                    {
                        "name": name,
                        "fret_row": row,
                        "start_col0": col - 1,  # 0-indexed: フレット1列目の列
                        "grid_row0_top": grid_row1_top - 1,  # 0-indexed top row
                        "start_fret": start_fret,  # このグリッドの最初の列が実際は何フレットか
                    }
                )
    return blocks


def shapes_in_block(shapes, block):
    """ブロックのグリッド範囲(0-indexed col/row)に属する図形を抜き出し、
    (string_index 1-4, fret_index 1-5, type) に変換する。"""
    col0 = block["start_col0"]  # 0-indexed col of フレット1
    row0 = block["grid_row0_top"]  # 0-indexed top row of grid (弦1の上端境界線)

    results = []
    for s in shapes:
        fret = s["from_col"] - col0 + 1  # 1,2,3,4,5
        if not (1 <= fret <= NUM_FRETS):
            continue
        if s["type"] == "flowChartTerminator":
            # セーハ: from_row~to_row が複数弦にわたる
            str_from = s["from_row"] - row0 + 1
            str_to = s["to_row"] - row0
            if str_from < 1 or str_to > NUM_STRINGS:
                continue
            results.append({"type": "bar", "fret": fret, "string_from": str_from, "string_to": str_to})
        else:
            string_idx = s["from_row"] - row0 + 1
            if not (1 <= string_idx <= NUM_STRINGS):
                continue
            results.append({"type": "dot", "fret": fret, "string": string_idx})
    return results


# ----------------------------------------------------------------------------
# 4. 画像描画
# ----------------------------------------------------------------------------
def safe_filename(name):
    # ファイル名に使えない文字を置換 (# / は特に注意)
    replacements = {
        "#": "s", "♭": "b", "/": "_on_", "\\": "_",
        "?": "", "*": "", ":": "-", '"': "", "<": "", ">": "", "|": "",
    }
    out = name
    for k, v in replacements.items():
        out = out.replace(k, v)
    return out


def draw_chord_image(name, shapes_data, start_fret=1):
    """
    フォントサイズとフレット1個分の幅は固定値を使う。コード名が長い場合は
    画像の横幅が自動的に広がる(画像サイズを揃えて文字を縮小する、という
    やり方はしない)。
    """
    WHITE = (255, 255, 255, 255)  # 不透明な白(線・文字・丸)
    is_open_position = start_fret == 1

    # --- 固定基準値(元画像 166x163, 5フレット表示の比率を基準に算出) ---
    FRET_W = 24           # フレット1個分の幅(px)。元画像相当(32.8px)より狭くする
    STRING_GAP = 13       # 弦と弦の間隔(px)
    TITLE_H = 34          # タイトル(コード名)エリアの高さ(px)
    SIDE_MARGIN = 4       # グリッド左右の余白(px)
    TOP_PAD = 0           # グリッド上端の追加余白
    BOTTOM_PAD_OPEN = 7   # 最下段の丸が見切れないための余白(オープンポジション)
    BOTTOM_PAD_HIGH = 20  # ハイポジション時は開始フレット番号の表示分を確保

    nut_w = 7 if is_open_position else 2
    line_w = 2

    grid_w = NUM_FRETS * FRET_W
    grid_h = (NUM_STRINGS - 1) * STRING_GAP
    bottom_pad = BOTTOM_PAD_OPEN if is_open_position else BOTTOM_PAD_HIGH

    grid_top = TITLE_H + TOP_PAD
    grid_left = nut_w + SIDE_MARGIN
    grid_right = grid_left + grid_w
    grid_bottom = grid_top + grid_h

    # --- フォント準備 ---
    font_path = None
    for fp in (
        "/usr/share/fonts/opentype/noto/NotoSansCJK-Black.ttc",
        "/usr/share/fonts/opentype/noto/NotoSansCJK-Bold.ttc",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
        "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf",
    ):
        if os.path.exists(fp):
            font_path = fp
            break
    font = (
        ImageFont.truetype(font_path, TITLE_FONT_SIZE, index=0)
        if font_path
        else ImageFont.load_default()
    )

    # --- 画像全体の幅を決める(タイトル文字幅 と グリッド幅 の大きい方) ---
    margin_left = 5
    margin_right = 8
    tmp_img = Image.new("RGBA", (10, 10))
    tmp_draw = ImageDraw.Draw(tmp_img)
    bbox = tmp_draw.textbbox((0, 0), name, font=font)
    text_w = bbox[2] - bbox[0]
    text_h = bbox[3] - bbox[1]

    content_w = max(grid_right + SIDE_MARGIN, margin_left + text_w + margin_right)
    W = int(content_w) + 1
    H = int(grid_bottom + bottom_pad) + 1

    img = Image.new("RGBA", (W, H), (0, 0, 0, 0))  # 背景は完全透明
    draw = ImageDraw.Draw(img)

    # --- タイトル(コード名)描画 ---
    draw.text((margin_left - bbox[0], (TITLE_H - text_h) / 2 - bbox[1]), name, fill=WHITE, font=font)

    # --- ナット(0フレット)の線(左端)。オープンポジションのみ太線にする。 ---
    draw.line([(grid_left, grid_top), (grid_left, grid_bottom)], fill=WHITE, width=nut_w)

    # --- 縦線(フレットの区切り線) ---
    for i in range(1, NUM_FRETS + 1):
        x = grid_left + i * FRET_W
        draw.line([(x, grid_top), (x, grid_bottom)], fill=WHITE, width=line_w)

    # --- 横線(弦): NUM_STRINGS本、等間隔 ---
    for j in range(NUM_STRINGS):
        y = grid_top + j * STRING_GAP
        draw.line([(grid_left, y), (grid_right, y)], fill=WHITE, width=line_w)

    # --- ハイポジションの場合、開始フレット番号を一番下の弦の下に表示する ---
    if not is_open_position:
        label_font_size = 16
        label_font = (
            ImageFont.truetype(font_path, label_font_size, index=0)
            if font_path
            else ImageFont.load_default()
        )
        label_text = str(start_fret)
        lbbox = draw.textbbox((0, 0), label_text, font=label_font)
        lw, lh = lbbox[2] - lbbox[0], lbbox[3] - lbbox[1]
        label_x = grid_left - lw / 2 - lbbox[0]
        label_y = grid_bottom + (bottom_pad - lh) / 2 - lbbox[1]
        draw.text((label_x, label_y), label_text, fill=WHITE, font=label_font)

    # --- セーハ・丸 (string番号 = 上から何本目の弦(線)か。1始まり) ---
    for sd in shapes_data:
        fret = sd["fret"]
        cx = grid_left + (fret - 0.5) * FRET_W
        if sd["type"] == "dot":
            string = sd["string"]
            cy = grid_top + (string - 1) * STRING_GAP
            r = min(FRET_W, STRING_GAP) * 0.32
            draw.ellipse([cx - r, cy - r, cx + r, cy + r], fill=WHITE)
        else:  # bar (セーハ): string_from弦からstring_to弦までを結ぶ縦長カプセル
            y_top = grid_top + (sd["string_from"] - 1) * STRING_GAP
            y_bot = grid_top + (sd["string_to"] - 1) * STRING_GAP
            r = FRET_W * 0.30
            pad = r * 0.5
            draw.rounded_rectangle(
                [cx - r, y_top - pad, cx + r, y_bot + pad], radius=r, fill=WHITE
            )

    return img


# ----------------------------------------------------------------------------
# メイン処理
# ----------------------------------------------------------------------------
def main():
    work_dir = extract_xlsx(XLSX_PATH, WORK_DIR)
    wb = load_workbook(XLSX_PATH, data_only=False)

    name_to_drawing = get_sheet_to_drawing_map(work_dir, wb.sheetnames)

    os.makedirs(OUT_DIR, exist_ok=True)

    manifest = []
    generated = 0
    skipped_sheets = []
    removed_exact_dups = []
    kept_variant_dups = []

    # シート内の同名コードをグルーピングし、重複を判定してから描画する
    for sheet_name in wb.sheetnames:
        if sheet_name not in name_to_drawing:
            skipped_sheets.append(sheet_name)
            continue
        ws = wb[sheet_name]
        drawing_file = name_to_drawing[sheet_name]
        drawing_path = os.path.join(work_dir, "xl", "drawings", drawing_file)
        if not os.path.exists(drawing_path):
            skipped_sheets.append(sheet_name)
            continue

        shapes = parse_shapes(drawing_path)
        blocks = find_chord_blocks(ws)

        # 同名コードごとにグルーピング
        by_name = {}
        for block in blocks:
            by_name.setdefault(block["name"], []).append(block)

        used_filenames_in_sheet = {}

        for name, block_list in by_name.items():
            # 各ブロックの押弦データを計算し、内容で重複排除する
            seen_signatures = []
            unique_blocks = []  # (block, shapes_data)
            for block in block_list:
                sd = shapes_in_block(shapes, block)
                signature = (
                    block["start_fret"],
                    tuple(sorted(tuple(sorted(d.items())) for d in sd)),
                )
                if signature in seen_signatures:
                    # 完全に同じ運指の重複 → 捨てる
                    removed_exact_dups.append(f"{sheet_name}シート「{name}」(row={block['fret_row']})")
                    continue
                seen_signatures.append(signature)
                unique_blocks.append((block, sd))

            if len(unique_blocks) > 1:
                kept_variant_dups.append(
                    f"{sheet_name}シート「{name}」({len(unique_blocks)}種類の異なる運指)"
                )

            for block, sd in unique_blocks:
                img = draw_chord_image(name, sd, start_fret=block["start_fret"])
                base_fname = safe_filename(name) + ".png"
                if base_fname in used_filenames_in_sheet:
                    used_filenames_in_sheet[base_fname] += 1
                    root, ext = os.path.splitext(base_fname)
                    fname = f"{root}_{used_filenames_in_sheet[base_fname]}{ext}"
                else:
                    used_filenames_in_sheet[base_fname] = 1
                    fname = base_fname

                out_path = os.path.join(OUT_DIR, fname)
                img.save(out_path)
                manifest.append({"name": name, "file": fname, "sheet": sheet_name})
                generated += 1

    with open(os.path.join(OUT_DIR, "..", "manifest.json"), "w", encoding="utf-8") as f:
        json.dump(manifest, f, ensure_ascii=False, indent=2)

    print(f"生成枚数: {generated}")
    if skipped_sheets:
        print(f"スキップしたシート(図形なし/未対応): {skipped_sheets}")
    if removed_exact_dups:
        print("\n[自動排除] 完全に同じ運指の重複コードを除外しました:")
        for d in removed_exact_dups:
            print("  - " + d)
    if kept_variant_dups:
        print("\n[情報] 同名だが運指が異なるコードが見つかったため、両方を残しました(要確認):")
        for d in kept_variant_dups:
            print("  - " + d)


if __name__ == "__main__":
    main()
